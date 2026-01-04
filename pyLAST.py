#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 21 23:07:37 2025
ptLAST LAST analysis functions package
Packs all of Ron Arad's functions into one package
With addtions and enhancements by Micha
@author: micha
"""
import pandas as pd
import numpy as np
import os, sys
from datetime import datetime, timedelta
from openpyxl import load_workbook
import json
import math
# import csv
import time
# import tomllib as tom
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
from sklearn.neighbors import KernelDensity
from scipy.optimize import minimize
from scipy.stats import gaussian_kde
from sklearn.preprocessing import StandardScaler
import matplotlib.ticker as ticker
import matplotlib.lines as mlines
# from collections import deque
# from io import StringIO
import subprocess
# import argparse
import seaborn as sns
# import clickhouse_connect
from clickhouse_driver import Client
import clickhouse_connect
# from pydantic import BaseModel
#------------------- end of imports
# A constant used in several funcions
t0 = pd.Timestamp("1970-01-01T00:00:00")
# The following colors are globally used for 4 traces (4 scopes) and for 10 mounts
colors = ['red', 'green', 'blue', 'purple']
colors_mounts = {"1": "red", "2": "green", "3": "blue", "4": "brown", "5": "pink", "6": "gray",    \
          "7": "purple", "8": "orange", "9": "olive", "10": "cyan"}  # per mount  
# constraints for fitting focus -temperature slopes 
constraints =(-20., 23., 16., 0.1)   #first 2 are the rigid slope limits, the third is the center and the 4th is the weight
fovdict ={'tl':[5,6,12],
      'tr':[18,24,23],
      'tc':[11,17],
      'dc':[8,14],
    'll': [3,4],
    'rr': [21,22],
    'dl': [1,2,7],
    'dr':[13,19,20],
    'ctr': [9,10,15,16]}
# class LastDatabase(BaseModel):
#     """
#     Defines a LAST database class
#     with properties: 
#         name - string
#         host = host ip string
#         port = port num integer
#         user = user name string
#         pw = password string
    
#     """
    
#     name:str = None
#     host:str = None
#     port:int = None
#     user:str = None
#     pw:str = None
#     operdb:str = None
class LastDatabase: 
    def __init__(self,dbcfg):
        self.name = dbcfg.name
        self.host = dbcfg.host
        self.port = dbcfg.port
        self.user = dbcfg.user
        self.pw = dbcfg.pw
        self.operdb = dbcfg.operdb
        
    def connect(self):
        if self.name == 'science':
            client = clickhouse_connect.get_client(host=self.host, port=self.port,
                                                   username=self.user, password=self.pw)
        else:
            client = Client(host=self.host, port=self.port, user=self.user,
                        password=self.pw, database= self.operdb)  
        return client
    # def __init__(self,name: str,host:str,port:int,user:str,pw:str):
    #     self.name = None
    #     self.host = None
    
class runconfig:   
    '''
    gernrates a run cofiguration object with all the config file sections.
    See the toml config file
    '''
    # localrun:bool = None
    # datbase:LastDatabase = None
    # input_path:str = None
    # output_path:str = None
    # database_path:str = None
    # Ndays:int = None
    # Nshow:int = None
    # startdate:str = None
    # enddate:str = None
    # fraction_to_read:float = None
    
    def __init__(self, cfg):
        for key, value in cfg.items():
            if isinstance(value, dict):
                # Recursively convert nested dictionaries to objects
                setattr(self, key, runconfig(value))
            elif isinstance(value, list):
                # Handle lists: convert dicts within lists to objects
                new_list = []
                for item in value:
                    if isinstance(item, dict):
                        new_list.append(runconfig(item))
                    else:
                        new_list.append(item)
                setattr(self, key, new_list)
            else:
                setattr(self, key, value)
    
class DualLogger:
    '''This class serves to log simultaneously to console + file'''
    def __init__(self, filename):
        self.terminal_stdout = sys.stdout
        self.terminal_stderr = sys.stderr
        self.log = open(filename, 'w', encoding='utf-8')

        sys.stdout = self
        sys.stderr = self

    def write(self, message):
        self.terminal_stdout.write(message)
        self.log.write(message)

    def flush(self):
        self.terminal_stdout.flush()
        self.log.flush()

    def close(self):
        sys.stdout = self.terminal_stdout
        sys.stderr = self.terminal_stderr
        self.log.close()
   

        
def generate_time_span_str(Ndays:int,Nshow:int,startdate:str=None,enddate:str=None) -> str:
    '''
    Generate a string giving the dates of the time span of the queries based on ther inputs.
    parameters:
        Ndays = number of days before today an optional start date for the span
        Nshow = the duuration in integer days from the start date
        startdate = a simple string in the format dd/mm/yy for the start date
        enddate = a single string in the format dd/mm/yy for the end date
    returns:
        time_span_str = a string describibg the time span in the format yymmdd_yymmdd 
        where the first part is the start date and the second is the end date  
        time span tuple (start-datedd-mm-yy,end-date dd-mm-yy)
    '''
    if startdate:
        start_date = datetime.strptime(startdate,'%d/%m/%y')
    elif Ndays:
        start_date = datetime.now() - timedelta(days=Ndays)
        start_date = start_date.replace(hour=12, minute=0, second=0, microsecond=0)
    else:
        raise Exception('Error: either start date or N_days must be given')
    
    if enddate:
        end_date = datetime.strptime(enddate,'%d/%m/%y')
    elif Nshow:
        end_date = start_date + timedelta(days=Nshow)
    else:
        raise Exception('Error: either start date or N_days must be given')
        
   
    # Compute end date (start_date + N_read days)
    # Format for ClickHouse (YYYY-MM-DD HH:MM:SS)
    start_str = start_date.strftime('%Y%m%d')
    end_str   = end_date.strftime('%Y%m%d')
    time_span_str = start_str + '_' + end_str
    return  ( time_span_str,start_date.strftime('%d/%m/%Y'),end_date.strftime('%d/%m/%Y'))

def save_df_to_csv(dfname:str,outdir:str, time_span_stamp:str,df:pd.DataFrame=None) -> str:
    '''
    Saves a pandas dataframe to a csv file in ther database dir
    uses the timespan string ti tag the file
    if the dataframe is not given the function generates the file name only 
    Parameters
    ----------
    
    dfname = a string with the data frame name
    outdir= the path of the data base dir 
    time_span_stamp : TYPE string
        DESCRIPTION = a time span stamp that will be added to the file name
    index = a flag to save an indexed dataframe default= True
    df : TYPE pandas data frame
        DESCRIPTION = the data frame to be saved
        if df is None it just returns the csv filename
    Returns the file name
    -------
    None.

    '''
    csv_file_name = time_span_stamp+'_'+dfname+'.csv'
    if df is not None:
        df.to_csv(os.path.join(outdir,csv_file_name))
    return csv_file_name

def read_DB1( client, db_name:str, rediskey_prefix:str, extra:str=None, 
             startdate:str=None, enddate:str=None, N_days:int =1, N_read:int=1):
    '''A general function for reading a given range of days from databse (dictionary with database details)
    looking for rediskey prefix. Note, the extra parameter is  mandatory and
    is used to add another condition
    parameters:
        client = a clickhouse database client opened with the database.connect method
        dB_name is either operation_strings or operation_numbers
        rediskey prefix - 
        examples of rediskey_prefix are: 'unitCS.set.FocusData:', 'XerxesMountBinary.get.Dec', 
        'XerxesMountBinary.get.Status'. An example of extra is: "value LIKE 'tracking'",
        "value like '%LoopCompleted%:%true%' "   
        extra = extra condition = query = query AND extra condition
        startdate = start date for the database query if None N_days and N_read are used
        enddate = end date for the database query if None N_days and N_read are used
        N_days is the number of days before present to search for (default = 1)
        N_read limits the search to N_read days after the initial day, set to -1 for all (default = 1)
        '''
    # if database.name == 'LAST_0':
    #     # LAST_0 as client using clickhouse_connect
    #     client = clickhouse_connect.get_client(host='10.23.1.25', port=8123, \
    #              username='last_user', password='physics', database='observatory_operation')        
    # elif database.name == 'euclid':
    #     # euclid as client using clickhouse_driver
    #     client = Client(host='euclid', port=9000, \
    #              user='last_user', password='physics', database='observatory_operation') 
    if N_read == -1: N_read=N_days
    query_string = build_range_query1( db_name, rediskey_prefix, extra, startdate, enddate, N_days, N_read)
    print(query_string)

    # if database.name == 'last0':    
    #     result = client.query(query_string)
    #     df = pd.DataFrame(result.result_rows, columns=['rediskey', 'time', 'value']).set_index('rediskey')
    # elif database.name == 'euclid':
    result = client.execute(query_string)
    df = pd.DataFrame(result, columns=["rediskey", "time", "value"]).set_index("rediskey")
        
    # Convert to DataFrame
    if df.empty:
        print('There is NO data in %s during the requested interval - stopping'% rediskey_prefix)
        sys.exit()  
    else: 
        print('loaded %d items'%len(df))
        print('Here is the first line:\n',df.tail(1))
        print('\n')
    return df

def read_visitDB( client, startdate:str=None, enddate:str=None, N_days:int=1, N_read:int=1)-> pd.DataFrame:
    '''A function for reading a given range of days from visit.images database 
    parameters:
        client = a clickhouse database client opened with the database.connect method
        startdate = start date (dd/mm/yy) for the database query if None N_days and N_read are used
        enddate = end date  (dd/mm/yy) for the database query if None N_days and N_read are used
        N_days is the number of days before present to search for (default = 1)
        N_read limits the search to N_read days after the initial day, set to -1 for all (default = 1)
        '''
    (start_str,end_str) = generate_date_range_str(startdate, enddate, N_days, N_read)
    query_string = f'''SELECT dateobs,mountnum,camnum,ra,dec,cropid,fwhm,med_a,med_b,med_th,airmass
                       FROM last.visit_images 
                       WHERE dateobs > '{start_str}'
                       AND dateobs < '{end_str}' '''
    print(query_string)
    df = client.query_df(query_string)
    if df.empty:
        print('There is NO data in during the requested interval - stopping')
        sys.exit()  
    else: 
        print('loaded %d items'%len(df))
        print('Here is the first line:\n',df.tail(1))
        print('\n')
    return df   
def telmap(cd:float)-> float:
    """
    Parameters
    ----------
    cd : float
        vector of 24 values of crop image data ordered from 1 to 24
        missing values in the vector  are replaced by nans

    Returns 
    -------
    telmap : float a 6x4 matrix of the crop data arranged as a 2d image.

    """
    telmap = np.zeros([6,4])
    cd = cd.astype(float)
    telmap = np.array(
             [[cd[5],cd[11],cd[17],cd[23]],
              [cd[4],cd[10],cd[16],cd[22]],
              [cd[3],cd[9],cd[15],cd[21]],
              [cd[2],cd[8], cd[14],cd[20]],
              [cd[1],cd[7],cd[13],cd[19]],
              [cd[0],cd[6],cd[12],cd[18]]])
    return telmap

def plot_mount_telescope_maps(mountnum:int,vals:float,
                              property_name:tuple, time_span_stamp:tuple,
                              outdir:str) -> plt.Figure:
    '''
    Plots 4 telescope maps of  a mount with annotations 
    of the values of vals and their stds
    Parameters:
        mountnum = the mount number
        vals = a list of 4 vectors (one for each telescope)
        of 24 float values extracted from the 24 crop images
        property_name[0],[1] = string the mapped property and its units
        time_span = a tuple of strings giving the time range of the data
        time_span_str(0,1) = start date:str, end date:str
        outdir = output directory to save the plots
    Returns
    -------
    figure with 2x2 subplots of telescope maps.

    '''
    
    mapfig,axs = plt.subplots(2,2,figsize = (12,10))
    axs = axs.flatten()
    for i, ax in enumerate(axs): # loop on telescopes
        valsmap = telmap(vals[i])
        # svalsmap = telmap(svals[i])
        sns.heatmap(valsmap,
                    vmin = np.nanmin(np.array(vals)),
                    vmax = np.nanmax(np.array(vals)),
                    annot=True,
                    fmt = ".1f",
                    cmap = 'inferno',
                    ax= ax,
                    cbar = False,
                    square = True
                    )
        ax.set_title('Telscope %d %s %s'%((i+1),property_name[0],property_name[1]))
    mapfig.suptitle(f"Mount {mountnum:d} {time_span_stamp[1]}-{time_span_stamp[2]} {property_name[0]}")
    plt.tight_layout()
    mapfig_filename = 'mount_%d_%s_%s_telmap.png'%(mountnum,time_span_stamp[0],property_name[0])
    mapfig.savefig(os.path.join(outdir,mapfig_filename))
    plt.show() # for debug
    return mapfig
    
    

def generate_date_range_str(startdate:str,enddate:str,ndays:int,nread:int)->tuple:
    if nread == -1: nread=ndays
    if startdate:
        start_date = datetime.strptime(startdate,'%d/%m/%y').replace(hour=12, minute=0, second=0, microsecond=0)
    elif ndays:
        start_date = datetime.now() - timedelta(days=ndays)
        start_date = start_date.replace(hour=12, minute=0, second=0, microsecond=0)
    else:
        raise Exception('Error: either start date or N_days must be given')
    
    if enddate:
        end_date = datetime.strptime(enddate,'%d/%m/%y').replace(hour=12, minute=0, second=0, microsecond=0)
    elif nread:
        end_date = start_date + timedelta(days=nread)
    else:
        raise Exception('Error: either start date or N_days must be given')
        
    start_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
    end_str   = end_date.strftime('%Y-%m-%d %H:%M:%S')
    date_range_str = (start_str,end_str)
    return date_range_str


def build_range_query1( table:str, rediskey_prefix:str, extra_condition=None,
                              startdate:str=None, enddate:str=None, N_days:int=1, N_read:int=1) -> str:
    """ Build a ClickHouse SQL query for a given time range.
    N_days : int       Number of days ago to start (start at 12:00:00 that day).
    N_read : int       Number of days to include in the query range.
    table : str        Name of the ClickHouse table (e.g., 'operation_strings').
    rediskey_prefix : str  The prefix string for startsWith(rediskey, ...).
    extra_condition (str, optional): Additional SQL condition to append to WHERE
    Returns  query : str    SQL query string. 
    start_date : string  start date dd/mm/yy  (start at 12:00:00 that day).
    end_date : string end_date dd/mm/yy (ends at 12:00 that day)
    """
    # Compute start date (N days ago at noon)
    # start_date = datetime.now() - timedelta(days=N_days)
    if startdate:
        start_date = datetime.strptime(startdate,'%d/%m/%y').replace(hour=12, minute=0, second=0, microsecond=0)
    elif N_days:
        start_date = datetime.now() - timedelta(days=N_days)
        start_date = start_date.replace(hour=12, minute=0, second=0, microsecond=0)
    else:
        raise Exception('Error: either start date or N_days must be given')
    
    if enddate:
        end_date = datetime.strptime(enddate,'%d/%m/%y').replace(hour=12, minute=0, second=0, microsecond=0)
    elif N_read:
        end_date = start_date + timedelta(days=N_read)
    else:
        raise Exception('Error: either start date or N_days must be given')
        
   
    # Compute end date (start_date + N_read days)
    # Format for ClickHouse (YYYY-MM-DD HH:MM:SS)
    start_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
    end_str   = end_date.strftime('%Y-%m-%d %H:%M:%S')

    # Build query
    query = f"""
        SELECT * FROM {table}
        WHERE startsWith(rediskey, '{rediskey_prefix}')
          AND time > '{start_str}'
          AND time < '{end_str}'    """
    # add extra condition if provided
    if extra_condition:
        query += f" AND {extra_condition}"   
    # always order by time descding
    query += " ORDER BY time DESC "         
    return query.strip()  
  


def filter_columns_by_nan_or_empty_lists(df: pd.DataFrame, frac_nans=0.1):
    """ Removes columns that have more than frac NaNs or empty lists.
    Assumes that each column contains only NaNs or only empty lists, not both.
    Parameters: frac: Maximum allowed fraction NaNs = N_NaNs/ len(df)
    Returns:    tuple: (filtered DataFrame, list of removed column names) """
    
    N = len(df)
    threshold = int(N * frac_nans)
    cols_to_remove = []
    for col in df.columns:
        col_data = df[col]
        
        if col_data.dtype == 'O':  # likely object type
            # Check for empty lists
            count_empty_lists = col_data.apply(lambda x: isinstance(x, list) and len(x) == 0).sum()
            if count_empty_lists > threshold:
                cols_to_remove.append(col)
                continue
        # Check for NaNs (for any dtype)
        count_nans = col_data.isna().sum()
        if count_nans > threshold:
            cols_to_remove.append(col)

    filtered_df = df.drop(columns=cols_to_remove)
    return filtered_df, cols_to_remove


def filter_columns_by_nan(df: pd.DataFrame, N=1000):
    """ Removes columns from a DataFrame that have more than N NaN values.
    Parameters:
        df (pd.DataFrame): The input DataFrame.
        N (int): The maximum allowed number of NaNs per column.
    Returns: tuple: (filtered DataFrame, list of removed column names) """
    
    # Find columns where the number of NaNs is greater than N
    cols_to_remove = df.columns[df.isna().sum() > N].tolist()
    
    # Drop those columns
    filtered_df = df.drop(columns=cols_to_remove)

    return filtered_df, cols_to_remove


def julian_to_ddmm(jd, fmt='%d-%m-%y'):
    """ Convert Julian Day number(s) to formatted date string(s).
    Parameters:
        jd (float or array-like): Julian Day(s)
        fmt (str): datetime.strftime format string (default '%d-%m')
    Returns: str or pd.Series: formatted date string(s)   """    
    # Convert input to numpy array for uniformity
    # 
    jd_arr = np.atleast_1d(jd)
    ts = pd.to_datetime((jd_arr - 2440587.5) * 86400, unit='s', utc=True)
    # Format dates as strings
    formatted = ts.strftime(fmt)
    # If input was scalar, return single string, else return Series
    if np.isscalar(jd):
        return formatted[0]
    else:
        return pd.Series(formatted, index=None)    


def filter_1hr_after_start(df, start_indexes, time_to_filter=0.0417):
    '''Filters some of the rows based on the 'TimeStarted' column. For each start index
    a portion corresponding to (time-to_filter, default= 1Hr) is removed from df. 
    This is meant for filtering the early hours of twilight in each day'''
    
    mask = pd.Series(True, index=df.index)  # start with all True (keep everything)
    col_idx = df.columns.get_loc('TimeStarted')
            
    for i in range(len(start_indexes)):
        start_time = df.iloc[start_indexes[i], col_idx]
        end_time = start_time + time_to_filter
    
        # Mask out the 1-hour region after each start_time
        mask &= ~((df['TimeStarted'] >= start_time) & (df['TimeStarted'] < end_time))
    
    # Apply the mask to keep only non-excluded rows
    return df[mask]


def linear_fit(x,y,name,fit_subtract_offset):
    '''This is the basic linear-fit function used by all the other functions
    the text it returns is for adding to the plots and is rounded accordingly
    It returns the predicted y values (line), the fit quality R2, the fit parameters
    m and b and the text for the plot'''    
    
    model = LinearRegression(fit_intercept=fit_subtract_offset)
    model.fit(x, y)
    y_pred = model.predict(x)
    R2 = model.score(x, y)
    m = model.coef_[0].item()
    b = model.intercept_.item() if fit_subtract_offset else 0.0
    R2_n = num_round(R2)
    m_n = num_round(m)
    b_n = num_round(b)
    # Format annotation
    name=str(name)
    text = f'{name}:\n$R^2$ = {R2:.{R2_n}f}\nm = {m:.{m_n}f}\nb = {b:.{b_n}f}'  
    return y_pred, R2, m, b, text


def linear_fit_intercept_only(x,y, name, slope_fixed):
    '''This is a simple calculation when we only want the intercept. 
    It si similr to the linear fit function and returns similar outputs'''
    
    # Compute intercept: mean(y) - m * mean(x)
    b = np.mean(y) - slope_fixed * np.mean(x)
    m = slope_fixed
    # Predictions
    y_pred = slope_fixed * x + b
    
    # R²
    ss_res = np.sum((y - y_pred.flatten())**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    R2 = 1 - ss_res/ss_tot
    b_print = num_round_2(b)
    R2_print = num_round_2(R2)
    print(f"Slope: {slope_fixed}, Intercept: {b_print}, R²: {R2_print}")
    R2_n = num_round(R2)
    m_n = num_round(m)
    b_n = num_round(b)
    # Format annotation
    name=str(name)
    text = f'{name}:\n$R^2$ = {R2:.{R2_n}f}\nm = {m:.{m_n}f}\nb = {b:.{b_n}f}'  
    return y_pred, R2, m, b, text

def compute_kde_weights(x, y, bandwidth=1.0):
    """ Compute KDE weights for points (x,y). Returns normalized weights in [0,1] """
    
    #xy = np.vstack([x, y]).T
    xy = np.column_stack([x, y])
    scaler = StandardScaler()
    xy_scaled = scaler.fit_transform(xy)  # now both axes have std=1
    #xy = np.column_stack([x, y])  # keep original units

    kde = KernelDensity(kernel='gaussian', bandwidth=bandwidth).fit(xy_scaled)
    log_dens = kde.score_samples(xy_scaled)
    weights = np.exp(log_dens)
    # Normalize so max = 1
    #weights /= np.max(weights)
    weights /= np.sum(weights)
    return weights

def fit_constrained_linear(x, y, name, slope_range=[13, 23], slope_center=18, slope_weight=1.0, bandwidth=1.0):
    """ Fit y = m*x + b with: slope constrained in [slope_range] + penalty favoring slope
    near slope_center slope constrained in [slope_range] penalty favoring slope near 
    slope_center KDE weights to emphasize dense regions     """
       
    x = np.asarray(x).ravel()
    y = np.asarray(y).ravel()
    
    # Compute KDE weights
    w = compute_kde_weights(x, y, bandwidth=bandwidth)

    def objective(params):
        m, b = params
        y_pred = m * x + b
        residuals = y - y_pred
        ssr = np.sum(w * residuals**2)
        # Penalty term: quadratic distance from slope_center
        penalty = slope_weight * (m - slope_center)**2
        return ssr + penalty

    # Initial guess using ordinary least squares
    m0, b0 = np.polyfit(x, y, 1)

    bounds = [(slope_range[0], slope_range[1]), (None, None)]

    result = minimize(objective, x0=[m0, b0], bounds=bounds)

    if not result.success:
        # Fallback: clamp slope to nearest bound and refit intercept
        if m0 < slope_range[0]:
            m = slope_range[0]
        elif m0 > slope_range[1]:
            m = slope_range[1]
        else:
            # if polyfit slope inside range but optimizer failed, use it directly
            m = m0
        # Refit intercept with chosen slope
        b = np.mean(y - m * x)
    else:
        m_fit, b_fit = result.x

    m, b = result.x
    y_pred = m * x + b

    # R² calculation
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    R2 = 1 - ss_res / ss_tot if ss_tot != 0 else 0.0
    R2_n = num_round(R2)
    m_n = num_round(m)
    b_n = num_round(b)
    name=str(name)
    text = f'{name}:\n$R^2$ = {R2:.{R2_n}f}\nm = {m:.{m_n}f}\nb = {b:.{b_n}f}'  
    # plt.scatter(x, y, c=w, cmap="viridis")
    # plt.plot(x, y_pred, "r")
    # plt.colorbar(label="KDE weight")
    # plt.show()

    return y_pred, R2, m, b, text


def num_round(x, sig_digits=3):
    ''' Estimate the optimal number of decimal places (n) for formatting a float x 
    in fixed-point notation ('.nf'), such that approximately 'sig_digits' 
    Initially it deals with 'NaN' and only then proceeds to yield significant 
    digits are preserved.'''
    
    if x == 'NaN': 
        return 'NaN'    
    else:
        if x == 0:
            return 0  # zero is special
        if isinstance(x, float):
            abs_x = abs(x)
            log10_x = math.log10(abs_x)
            n = sig_digits - 1 - math.floor(log10_x)
        else:
            return 0
        return max(0, n)


def num_round_2(x, sig_digits=3):
    if x == 'NaN': 
        return 'NaN'    
    else:
        N = num_round(x, sig_digits)
        return round(x, N)


def filter_N_days(df, n_days, n_show = -1 ):
    """ Filters the df to include only rows from the last `n_days`
    based on the 'TimeStarted' column (in units of days). The optional parameter
    N_show can limit the viewing so that only the first days are shown """
    
    if n_show == -1: n_show=n_days
    if 'TimeStarted' not in df.columns:
        raise ValueError("df must contain a 'TimeStarted' column.")
    today = datetime.today()
    today_noon = datetime(today.year, today.month, today.day, 12, 0, 0)
    today_noon_julian = today_noon.toordinal() + 1721425
    start_time = today_noon_julian - n_days
    end_time = start_time + n_show
    #print('date_range_crop:', julian_to_ddmm(start_time,fmt='%d-%m-%H-%M'), julian_to_ddmm(end_time,fmt='%d-%m-%H-%M'))
    return df[(df['TimeStarted'] >= start_time) & (df['TimeStarted'] < end_time)]


def filter_groups_N_days(df_groups, n_days, n_show = -1 ):
    """ Applies filter_N_days to each DataFrame in a list of DataFrames.
    Returns: List of filtered DataFrames  """

    return [filter_N_days(df, n_days, n_show) for df in df_groups]


def plot_saving(output_directory, file_name, label):
    '''A general function for saving the plotted images to a folder called label.
    the label tell the program what group of graphs this belongs to and creates a 
    subfolder in the main output directory. The filename (without the extension)
    is then used to generate a full filename and save it'''
    
    folder_name = f"{label}"     
    folder_path = os.path.join(output_directory, folder_name)
    os.makedirs(folder_path, exist_ok=True)   # Create the output directory if it doesn't exist
    filtered_file_name = f"{file_name}.png"
    full_file_path = os.path.join(folder_path, filtered_file_name) 
    plt.savefig(full_file_path, format='png', dpi=300, bbox_inches='tight')
    return full_file_path


def highlight_bad_R2(column):
    '''Mark column in orange if v < threshold'''
    
    # used for applying to the excel output of the fit parameters
    R2_threshold = 0.3
    return ['background-color: orange' if (v < R2_threshold or v > 1/R2_threshold) else '' for v in column]


def highlight_bad_slope(column):
    '''Mark column in red if slope is outside range 10-25'''
    
    # used for applying to the excel output of the fit parameters
    slope_min = 10
    slope_max = 25
    return ['background-color: red' if (v < slope_min or v > slope_max) else '' for v in column]
    

def write_append_cols_excel(Fit_results,output_directory):
    '''Exports the output to excel. When repeated, in order
    to append columns next to previously placed columns, the initial export is
    first imported and then the startcol is adjusted after the existing data'''
    
    out = os.path.join(output_directory,(file_short + '_Fit_results.xlsx'))
    Fit_results = pd.DataFrame(Fit_results)
    #styled = Fit_results.style.map(highlight_bad_R2, subset=['R2_score'])
    #styled = Fit_results.style.map(highlight_bad_slope, subset=['slope'])
    styled = (Fit_results.style.apply(highlight_bad_R2, subset=['R2_score'])
              .apply(highlight_bad_slope, subset=['slope'])  )
    
    # Check if file exists
    if os.path.exists(out):
        # Load existing workbook
        wb = load_workbook(out)
        ws = wb.active  
        start_col = ws.max_column  # append after last existing column
        existing_df = pd.read_excel(out, sheet_name=ws.title)
        
        with pd.ExcelWriter(out, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            styled.to_excel(writer, sheet_name=ws.title, index=False, header=True, startcol=start_col)
            # Compute ratio between matching columns (only for numeric ones)
        common_cols = Fit_results.columns.intersection(existing_df.columns)
        
    else:
        # File doesn't exist: create new file with styled DataFrame
        styled.to_excel(out, engine='openpyxl')


def keep_2_elements_only(s):
    '''removes centeral number only if it is = 1
    This function is used to remove redundant '1' in the mount.telescope numbers '''

    parts = s.split('.')
    if len(parts) == 3 and parts[1]=='1':
        return f"{parts[0]}.{parts[2]}"
    return s 


def load_FWHM_csv(input_file, fraction_to_read, start_reading_at_frac):
    '''This is the basic import and initial basic manipulation of the df it reads, splits the json part and
    adds a julian date + day counter'''
    
    # Step 1: Read and parse JSON
    #df = pd.read_csv(input_file, sep=',', header=None, index_col=0, quotechar='"', engine='python')
    #instead of normal read, use deque to read last part and only fraction
    print()
    line_count = fast_line_count(input_file)
    start = time.time()
    N = int(fraction_to_read * line_count)
    offset = int(start_reading_at_frac * line_count)
    df = pd.read_csv(input_file, sep=',',skiprows=offset+1, nrows=N ,header=None, index_col=0, quotechar='"', engine='python')      
    end = time.time()
    print(f"{input_file} Elapsed time: {end - start:.4f} seconds")
    
    # The next 4 lines are for Reading the last N lines using deque and then join and read into df
    # with open(input_file, 'r', encoding='utf-8') as f:
    #     last_N_lines = deque(f, maxlen=N)
    # data = StringIO(''.join(last_N_lines))
    # df = pd.read_csv(data, sep=',', header=None, index_col=0, quotechar='"', engine='python')
        
    df.columns = ['time', 'value']
    print('loaded ', input_file.split('/')[-1], ' read last ', N, ' lines')
    first_row = df.iloc[0]
    last_row = df.iloc[-1]
    print(first_row['time'].split(' ')[0], ' until ', last_row['time'].split(' ')[0])
    return df
    
    
def basic_processing_FWHM(df):
    '''Basic processing includes preparing index col to only include mount.scope, replace _ by .
    THen, open the json and Create a col called HH-MM-DD-mm and (if missing) add a column TimeStarted
    returns the input df sorted by TimeStarted. Also, the minor and major columns are mult. by 2.355.
    Finally, empty fit rows are removed and only succesful fits maintained'''
    
    df['value'] = df['value'].apply(json.loads)
    df.index = df.index.str.split(':').str[-1]
    #repair some of the errors in the index
    df.index = df.index.astype(str).str.replace('_', '.', regex=False)
    #sometimes the mount.telescope numbers have a '.1' between them, this removes it
    df.index = df.index.map(keep_2_elements_only)
    
    
    # Step 2: Normalize top-level fields (excluding ResTable)
    top_level = pd.json_normalize(df['value'])
    top_level.index = df.index
    top_level_no_restable = top_level.drop(columns='ResTable', errors='ignore')
    #print('original size', top_level_no_restable.shape)
    # Step 3: Extract and flatten ResTable into wide format
    restables = []
    
    # Loop over each row to flatten ResTable into columns
    '''Currently, don't want  the restable so am not adding it'''
    require_also_ResTable = False
    if require_also_ResTable:
        for idx, row in top_level.iterrows():
            row_data = {}
            restable = row.get('ResTable', None)
        
            # Make sure ResTable is a list of dicts
            if isinstance(restable, list):
                for i, item in enumerate(restable):
                    if isinstance(item, dict):
                        for key, value in item.items():
                            col_name = f'ResTable_{i}.{key}'
                            row_data[col_name] = value
            # Build a Series for the row
            row_series = pd.Series(row_data, name=idx)
            restables.append(row_series)
        
        # Combine ResTable expansions into a DataFrame
        restable_df = pd.DataFrame(restables)
    
    # Step 4: Combine everything and add a time column
    if require_also_ResTable:
        final_df = pd.concat([top_level_no_restable, restable_df], axis=1)
    else:
        final_df = pd.concat([df['time'], top_level_no_restable], axis=1)
    # make sure the 'time' column is of type datetime
    if 'time' in df.columns:
        df["time"] = pd.to_datetime(df["time"], errors="coerce")
    #Create a column in HH-dd-mm notation for the start time
    if 'TimeEnded' in final_df.columns:
        final_df["HH-MM-DD-mm"] = final_df["TimeEnded"].apply(lambda jd: julian_to_ddmm(jd, fmt='%H_%M_%d_%m'))
    # if TimeStarted coloumn is missing, generate it from time    
    elif 'time' in final_df.columns and 'TimeStarted' not in final_df.columns:
        final_df["HH-MM-DD-mm"] = pd.to_datetime(final_df['time']).dt.strftime('%H_%M_%d_%m')
        final_df['time'] = pd.to_datetime(final_df['time'], errors='coerce')   
        # subtracting t0 (1970) to make a time-diff and converting to days and adding (1970-jd) to convert to jd
        final_df['TimeStarted'] = (final_df['time'] - t0).dt.total_seconds() / (24 * 3600) + 2440587.5

    '''for FWHM data need to convert minor and major columns by multiplying by 2*sqrt(ln(4)) = 2.355'''
    if 'minor' in final_df.columns:
        final_df['minor'] = final_df['minor'] * 2.355 
    if 'major' in final_df.columns:
        final_df['major'] = final_df['major'] * 2.355     

    #Now we filter:
    #final_df = final_df[final_df["Status"] != "Fault"]
    if 'LoopCompleted' in final_df.columns:
        final_df = final_df[final_df["LoopCompleted"] == True]
    '''for focus we want to filter out bad fit attempts'''
    if 'FitParam' in df:
        df = [df["FitParam"] != [None, None, None]] # remove focus sesstions that did not produce a good Fit
        #df = [df['Status'] == ['Found.']]  keep only good "Found" focuses
    return final_df.sort_values(by='TimeStarted', ascending=True)


def load_tracking_csv(input_file, fraction_to_read, start_reading_at_frac):
    '''This is the basic import and initial basic manipulation of the tracking data 
    it reads adds a julian date + day counter'''
    
    # Step 1: Read and parse JSON
    #df = pd.read_csv(input_file, sep=',', header=None, index_col=0, quotechar='"', engine='python')
    #instead of normal read, use deque to read last part and only fraction
    print()
    line_count = fast_line_count(input_file)
    start = time.time()
    N = int(fraction_to_read * line_count)
    offset = int(start_reading_at_frac * line_count)
    df = pd.read_csv(input_file, sep=',',skiprows=offset, nrows=N ,header=None, index_col=0, quotechar='"', engine='python')      
    if 'time' in df.columns:
        df["time"] = pd.to_datetime(df["time"], errors="coerce")
    if 'value' in df.columns:
        df["value"] = pd.to_numeric(df["value"], errors="coerce")
    end = time.time()
    print(f"{input_file} Elapsed time: {end - start:.4f} seconds")

    # The next 4 linea are for reading the last N lines using deque and then join and read into df
    # with open(input_file, 'r', encoding='utf-8') as f:
    #     last_N_lines = deque(f, maxlen=N)
    # data = StringIO(''.join(last_N_lines))
    # df = pd.read_csv(data, sep=',', header=None, index_col=0, quotechar='"', engine='python')
       
    df.columns = ['time', 'value']
    print('loaded ', input_file.split('/')[-1], ' read last ', N, ' lines')
    first_row = df.iloc[1]
    last_row = df.iloc[-1]
    print(first_row['time'].split(' ')[0], ' until ', last_row['time'].split(' ')[0])
    return df
    
def basic_processing_tracking(df):   
    '''Basic processing includes preparing index col to only include mount.scope, replace _ by .
    Create a col called HH-MM-DD-mm and (if missing) add a column TimeStarted
    returns the input df sorted by TimeStarted'''
    
    df.index = df.index.str.split(':').str[-1]
    #repair some of the errors in the index
    df.index = df.index.astype(str).str.replace('_', '.', regex=False)
    df = df.iloc[1:].copy() # remove the first row since someimes from csv it includes headers
    final_df = df
    
    #Create a column in HH-dd-mm notation for the start time
    if 'TimeEnded' in final_df.columns:
        final_df["HH-MM-DD-mm"] = final_df["TimeEnded"].apply(lambda jd: julian_to_ddmm(jd, fmt='%H_%M_%d_%m'))
    # if TimeStarted coloumn is missing, generate it from time    
    elif 'time' in final_df.columns and 'TimeStarted' not in final_df.columns:
        final_df["HH-MM-DD-mm"] = pd.to_datetime(final_df['time']).dt.strftime('%H_%M_%d_%m')
        final_df['time'] = pd.to_datetime(final_df['time'], errors='coerce')   
        # subtracting t0 (1970) to make a time-diff and converting to days and adding (1970-jd) to convert to jd
        final_df['TimeStarted'] = (final_df['time'] - t0).dt.total_seconds() / (24 * 3600) + 2440587.5

    return final_df.sort_values(by='TimeStarted', ascending=True)

        
def separate_by_mount(df):    
    '''split into 10 parts for the different mounts. Use integer part of the index
    Create a list with 10 elements, each a df. It checks if the index column has 3 elements,
    it removes the extra '1' in the middle. It creates mount and scope and adjusted_hour columns '''
    df = df.copy()
    # df.set_index('rediskey')
    #remove rows that don't have the mount.scope correctly
    #df = df[df['mount'].str.match(r"^\d+\.\d+$")]
    #df = df[df.index.to_series().str.match(r"^\d+\.\d+$")]
    
    #remove rows that don't have the mount.scope correctly, check first row, then
    # if 3 elements, remove middle one, Then keep only rows with the correct format 
    first_index = df.index[0]
    if str(first_index).count('.') == 2:  # means 3 elements
        df.index = df.index.map(lambda x: f"{x.split('.')[0]}.{x.split('.')[-1]}")
        print("had to remove extra '1.'")
    #remove rows that don't have the mount.scope correctly
    ###df = df[df.index.to_series().str.match(r"^\d+\.\d+$")]
    df = df[df.index != ""].copy()
    int_part, frac_part = np.divmod(df.index.to_numpy(dtype=float), 1)
    # Assign mount (0–9) and scope (0–3)
    df['mount'] = (int_part).astype(int)
    df['scope'] = np.round(frac_part * 10).astype(int)
    df['int_index'] =  int_part.astype(int) 
    
    # add a column of matched time HH:MM looping beyond 24hrs
    df['hour'] = df['HH-MM-DD-mm'].str.split('_').str[0].astype(float) + df['HH-MM-DD-mm'].str.split('_').str[1].astype(float)/60.
    df['adjusted_hour'] = df['hour']
    df.loc[df['hour'] < 5, 'adjusted_hour'] += 24
    
    # Group by int_index
    groups = []
    # reference columns (take from first available df)
    ref_cols = df.columns.drop("int_index")  # all columns except int_index   
    for i in range(1, 11):  # 10 mounts starting at number 1
        group = df[df["int_index"] == i].drop(columns="int_index").reset_index(drop=True)
        if group.empty:
            # create an empty DataFrame with same columns
            group = pd.DataFrame(columns=ref_cols)
        groups.append(group)
    return groups


def tracking_windows(groups):     
    '''after separating by groups, for the tracking data we need start and finish windows
     of the tracking. We create both columns + a duration column and to his end, 
     we subtracte 15 s (1.75e-4 days) from the finish, to account for the slewing to the new powition. '''
    
    slewing_time = 1.75e-4 #1.75e-4
    tracking_windows = []
    #for each mount add start, finish and duration columns    
    for group in groups:
        group = group.copy()
        group['start'] = group['TimeStarted']       
        # Shift 'TimeStarted' column up and subtract adjustment for 'finish'
        group['finish'] = group['start'].shift(-1) - slewing_time
        group['duration'] = ( group['finish']-group['start']) *24 *3600
        tracking_windows.append(group)   
    return tracking_windows

        
def analyze_tracking(tracking_groups, RA_groups, Dec_groups, Az_groups, Alt_groups):
    '''for each mount we look at the tracking times and then extract from the 
    RA and Dec the values during those tracking windows only for long tracking
    (50-1200s), which correspond to observation windows. Also, we filter out
    observations starting after 17:00.
    For RA require std< 0.002 for Dec require std < 10-4. This should be OK as long
    as the mount is really tracking
    The same is done for the Alt, Az (without std filtering)
    We average the RA and Dec and get a standard deviation for each observation.
    The combined data are then returned in a new df called tracking_results'''
    tracking_results = []
    
    for mount, track in enumerate(tracking_groups):
        RA_group = RA_groups[mount]
        Dec_group = Dec_groups[mount]
        Az_group = Az_groups[mount]
        Alt_group = Alt_groups[mount]
        matches = []
        count = 0
        count2 = 0
        for _, row in track.iterrows():
            
            start = row['start']
            finish = row['finish']
            duration = row['duration']
            start_hour =row['adjusted_hour']
            # Skip if finish is NaN (e.g., last row)
            if pd.isna(finish):
                continue
            if duration > 50 and duration < 1200 and start_hour > 17:  
                
                # Select rows in Az_group where TimeStarted is within [start, finish]
                RA_matched_rows = RA_group[(RA_group['TimeStarted'] >= start) & (RA_group['TimeStarted'] <= finish)]
                # Find rows where 'value' cannot be converted to a number              
                RA_matched_rows['value'] = pd.to_numeric(RA_matched_rows['value'], errors='coerce')
                if not RA_matched_rows.empty and RA_matched_rows['value'].std() < .002:
                    RA_median_val = RA_matched_rows['value'].median()
                    RA_std_val = RA_matched_rows['value'].std()
                else:
                    RA_median_val = 0
                    RA_std_val = 0
                    count+=1
                    
                    
                # Select rows in Alt_group where TimeStarted is within [start, finish]
                Dec_matched_rows = Dec_group[(Dec_group['TimeStarted'] >= start) & (Dec_group['TimeStarted'] <= finish)]
                Dec_matched_rows['value'] = pd.to_numeric(Dec_matched_rows['value'], errors='coerce')
                if not Dec_matched_rows.empty and Dec_matched_rows['value'].std() < 1e-4:
                    Dec_median_val = Dec_matched_rows['value'].median()
                    Dec_std_val = Dec_matched_rows['value'].std()
                else:
                    Dec_median_val = 0
                    Dec_std_val = 0
                    count2+=1
                    
                
                # Select rows in Az_group where TimeStarted is within [start, finish]
                Az_matched_rows = Az_group[(Az_group['TimeStarted'] >= start) & (Az_group['TimeStarted'] <= finish)]
                #if not Az_matched_rows.empty and Az_matched_rows['value'].std() < .002:
                Az_matched_rows['value'] = pd.to_numeric(Az_matched_rows['value'], errors='coerce')
                if not Az_matched_rows.empty :
                    Az_median_val = Az_matched_rows['value'].median()
                    Az_std_val = Az_matched_rows['value'].std()
                else:
                    Az_median_val = 0
                    Az_std_val = 0
                    count+=1
                    
                    
                # Select rows in Alt_group where TimeStarted is within [start, finish]
                Alt_matched_rows = Alt_group[(Alt_group['TimeStarted'] >= start) & (Alt_group['TimeStarted'] <= finish)]
                #if not Alt_matched_rows.empty and Alt_matched_rows['value'].std() < 1e-4:
                Alt_matched_rows['value'] = pd.to_numeric(Alt_matched_rows['value'], errors='coerce')                    
                if not Alt_matched_rows.empty :
                    Alt_median_val = Alt_matched_rows['value'].median()
                    Alt_std_val = Alt_matched_rows['value'].std()
                else:
                    Alt_median_val = 00
                    Alt_std_val = 0
                    count2+=1
                #store result (e.g., per-row summary)
                matches.append({
                    'HH-MM-DD-mm': row['HH-MM-DD-mm'],
                    'adjusted_hour': row['adjusted_hour'],
                    'start': start,
                    'finish': finish,
                    'duration': duration,
                    'RA_median': (RA_median_val + 180) % 360 - 180, # wraps the RA to be -180 up to +180
                    'RA_std': RA_std_val*3600 * np.cos(np.radians(Dec_median_val)),  #convert to arcsec and multiply by cos(Dec)
                    'Dec_median': Dec_median_val,
                    'Dec_std': Dec_std_val*3600,  #convert to arcsec
                    'Az_median': (Az_median_val + 180) % 360 - 180, # wraps the RA to be -180 up to +180
                    'Az_std': Az_std_val*3600 * np.cos(np.radians(Alt_median_val)),  #convert to arcsec and multiply by cos(Dec)
                    'Alt_median': Alt_median_val,
                    'Alt_std': Alt_std_val*3600  #convert to arcsec
                })
        #print('No matches found(mount,RA,Dec)', mount, count, count2) #This can be used to check for no matches between "tracking" and RA/ Dec data
        tracking_results.append(pd.DataFrame(matches))           

        #protect against empty groups by filling the column names manually
        new_cols = ['HH-MM-DD-mm', 'adjusted_hour', 'start', 'finish', 'duration', 
                    'RA_median', 'RA_std', 'Dec_median', 'Dec_std', 'Az_median', 
                    'Az_std', 'Alt_median', 'Alt_std']
        for i, df_group in enumerate(tracking_results):
            if df_group.empty:
                # create empty columns manually
                for col in new_cols:
                    df_group[col] = []  
                tracking_results[i] = df_group

    return tracking_results


def find_noon_rollover_indexes(df):
    """ Returns a list of row indexes where a new Julian day starts,
    assuming day changes at noon. df Must contain a 'TimeStarted' column in Julian Days.   
    Returns: list of indexes where a new day begins.  """
    
    if 'TimeStarted' not in df.columns:
        raise ValueError("DataFrame must contain 'TimeStarted' column")

    # Don't need to shift by 0.5, since my data are already 
    shifted_days = np.floor(df['TimeStarted'])
    # # Find where the day changes compared to the previous row
    day_change = shifted_days.diff().fillna(0) != 0.

    # Return the indexes where day change is True   
    rollover_indexes =  [i for i, change in enumerate(day_change) if change]
    
    return [0] + rollover_indexes #add the 0 index for convenience later in the plotting
        
def plot_bad_fit_fraction(df: pd.DataFrame, output_directory:str, instrument: str = "rediskey", status_col: str = "Status"):
    """
    For each unique value in the rediskey column, calculates the fraction of rows
    where the status does NOT contain "Found." and plots the result.

    Parameters:
        df (pd.DataFrame): Input DataFrame
        rediskey_col (str): Name of the column containing rediskeys
        status_col (str): Name of the column containing status messages
    """
    # Define a boolean mask for bad data (status does NOT contain 'Found.')
    is_bad = ~df[status_col].astype(str).str.contains(r"Found\.", regex=True, na=False)

    # Group by rediskey and calculate fraction of bad rows
    result = (
        df.assign(is_bad=is_bad)
          .groupby(df.index)["is_bad"]
          .mean()
          .sort_index()
    )
    # Plot
    plt.figure(figsize=(10, 5))
    result.plot(kind='bar', color='salmon', edgecolor='black')
    
    earliest = df["time"].min().strftime("%d_%m")
    latest = df["time"].max().strftime("%d_%m") 
    title = 'Fraction of Bad Focus by instrument ' + earliest + ' --- ' + latest 
    plt.title(title)
    plt.ylabel("Fraction of Bad Status")
    plt.xlabel("instrument")
    plt.xticks(rotation=90, ha='right')
    plt.ylim(0, 1)
    plt.grid(axis='y', linestyle='--', alpha=0.5)
    plt.tight_layout()
    filename = f'Fraction_Bad_focus {earliest}-{latest}'
    label = str('focus')
    plot_saving(output_directory, filename, label)
    plt.show()


def plot_bad_fraction_per_day(df, time_col='TimeStarted', status_col='Status'):
    """
    Plot fraction of bad fits per integer Julian day.
    Parameters:
        df: pandas DataFrame
        time_col: column with Julian day float values
        status_col: column with status strings
    """
    # Take integer part of Julian day as the day
    df['Day'] = df[time_col].astype(int)

    # Identify bad fits (Status NOT containing "Found.")
    is_bad = ~df[status_col].astype(str).str.contains(r"Found\.",
                                                      regex=True, na=False)
    df['is_bad'] = is_bad

    # Group by day and calculate fraction of bad fits
    daily_fraction = df.groupby('Day')['is_bad'].mean()

        # Prepare the plot
    fig, ax = plt.subplots(figsize=(12, 6))
    daily_fraction.plot(marker='o', linestyle='-', ax=ax)
    ax.set_title("Fraction of Bad Fits per Day")
    ax.set_xlabel("Julian Day")
    ax.set_ylabel("Fraction of Bad Fits")
    ax.set_ylim(0, 1)
    #ax.grid(True, linestyle='--', alpha=0.5)
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)


    # Add vertical grid lines on each integer day
    days = daily_fraction.index.values
    num_labels = min(5, len(days))
    label_days = np.linspace(days.min(), days.max(), num=num_labels, dtype=int)

    for day in label_days:
        # Add vertical line + dd-mm notation
        ax.axvline(day, color='gray', linestyle=':', alpha=0.5)
        ax.text(day, 0.8, julian_to_ddmm(day), rotation=90, ha='center', 
                va='bottom', fontsize=12, color='blue')

    plt.tight_layout()
    plt.show()

#TODO seperate plot from data processing


def plot_bestpos_vs_temp_by_mount(df: pd.DataFrame, output_directory: str,
                                  plots: dict, regular_fit: bool,
                                  x_axis: str, y_axis: str) -> pd.DataFrame:
    """ Plots y-axis data vs. x-axis data.
    It splits the data by mount number (int(indexes)) and plots nmounts (10) 
    subplots.
    Each subplot includes 4 scopes (based on fractional part of indexes).
    The y-limits are set by the 99th percentile of y_axis for that mount.
    Then, a linear fit is performed for each trace and plotted and saved.
    Specifically for BestPos the tick numbers are shifted to place assure all 4 
    scopes are in the same range with 150 tick spacing. The linear fit is 
    either simple,     or with constraints + center + weight to favor clusters,
    see the definition in Main.     Different symbols are given for the 
    first 12 days in a run. The first 2 hrs of observations are excluded
    (to reduce noise)
    parameters:
        df =
        output_directory = folder for saving the plots
        plots = a runconfig object with boolean flags to determine which plots to 
        plot
    """
    data_for_distribution_plot = []
    if y_axis == 'minor':
        dt = np.diff(df[x_axis])
        gap_indices = np.where(dt >= 0.05)[0] + 1
        start_indexes = np.insert(gap_indices, 0, 0)
        print(len(start_indexes), start_indexes)
        if len(start_indexes):
            col_idx = df.columns.get_loc('HH-MM-DD-mm')
            print(df.iloc[start_indexes, col_idx])
            df = filter_1hr_after_start(df, start_indexes, 0.083)
            # filter out 2 hrs
    if y_axis == 'BestPos':
        df = df[df['Status'] == 'Found.'].copy()
        # keep only focus data that is good
        print()
    medians_list = []   # this is for exporting the median values
    indexes = df.index.to_numpy(dtype=float)
    df = df.copy()
    int_part, frac_part = np.divmod(indexes, 1)
    # Assign mount (0–9) and scope (0–3)
    df['mount'] = int_part.astype(int)
    df['scope'] = np.round(frac_part * 10).astype(int)
    # Setup 10 mounts (0–9)
    fig, axes = plt.subplots(5, 2, figsize=(14, 20), sharex=True)
    axes = axes.flatten()

    if 'TimeStarted' in df.columns:
        latest_time = julian_to_ddmm(df['TimeStarted'].max())
        earliest_time = julian_to_ddmm(df['TimeStarted'].min())
    elif 'time' in df.columns:
        latest_time = df['time'].max()
        earliest_time = df['time'].min()

    title = '' + y_axis + ' vs. ' + x_axis + ' for: ' + \
        str(earliest_time) + ' --- ' + str(latest_time)

    # Prepare to collect fit results
    fit_results = []
    high_values = []
    for mount_num in sorted(df['mount'].unique()):
        ax = axes[mount_num-1]  # axes run from 0-9 and mount numbers are 1-10
        # Filter for this mount
        df_mount = df[df['mount'] == mount_num]

        # Compute initial y-axis limits for each mount
        median = df_mount[y_axis].median()
        lower_limit_mount = median
        upper_limit_mount = median

        # Plot each scope
        # this places the traces one above the other spaced by 150
        y_goal = [0, 34850, 34700, 34550, 34400]
        for scope_num in range(1, 5):  # keeps scope names 1-4
            df_scope = df_mount[df_mount['scope'] == scope_num]
            if df_scope.empty:
                print(f'mount {mount_num} scope {scope_num} is empty')
                std_int = 0
                continue
            # Calculate scope-specific thresholds
            median = df_scope[y_axis].median()
            std = df_scope[y_axis].std()
            q99 = df_scope[y_axis].quantile(0.99)
            range99 = abs(q99 - median)

            # Define lower and upper limits for filtering scope data and y-limit
            # taking 3x 99 percentile makes sure all points are kept except for
            #a crazy outlier
            lower_limit = median - 3*range99
            upper_limit = median + 3*range99

            # append a row
            if pd.isna(median):
                median_int = None  # or keep as np.nan
            else:
                median_int = int(median)
            if pd.isna(std):
                std_int = None  # or keep as np.nan
            else:
                std_int = int(std)
            medians_list.append({
                'mount': mount_num,
                'scope': scope_num,
                'median': median_int,
                'std': std_int})
            date_range = str(earliest_time) + ' --- ' + str(latest_time)
            try:
                if std_int > 10:
                    data_for_distribution_plot.append((list(df_scope[y_axis]),
                                                       median_int, std_int, 
                                                       int(mount_num), 
                                                       scope_num, date_range))
            except (NameError, TypeError): 
                # std_int is not defined or invalid (e.g., None)
                pass
            # Filter BestPos to within allowed range
            df_scope = df_scope[(df_scope[y_axis] >= lower_limit) &
                                (df_scope[y_axis] <= upper_limit)]
            '''Here can still split df_scope into separate days'''
            indexes = find_noon_rollover_indexes(df_scope)
            # print(f'\n mount {mount_num} scope{scope_num} has {indexes}')
            markers = ['o', '+', '^', 'D', 'P',
                       '*', 'x', 'v', '<', '>', '.', 's']
            x = df_scope[x_axis].to_numpy()
            # y = df_scope[y_axis].to_numpy() #absolute value of focus
            y = df_scope[y_axis].to_numpy()-median+y_goal[scope_num]

            # Remove NaNs
            valid = ~np.isnan(x).flatten() & ~np.isnan(y)
            x = x[valid].reshape(-1, 1)  # <== FIX: reshape here again
            y = y[valid]
            # print('valid pts:',len(y))
            indexes = indexes + [len(y)]

            # Linear fit
            if len(x) >= 2:
                if regular_fit:
                    '''For regular linear fit'''
                    y_pred, R2, m, b, text = linear_fit(x, y, 'col', True)
                elif not constraints:
                    '''For fixed slope and calculate only intercept use:'''
                    slope_fixed = 18
                    y_pred, R2, m, b, text = linear_fit_intercept_only(
                        x, y, 'col', slope_fixed)
                else:
                    y_pred, R2, m, b, text = fit_constrained_linear(x, y,'col',
                                           [constraints[0], constraints[1]], 
                                           constraints[2], constraints[3], 
                                           bandwidth=0.2)
                # print(f"mount {mount_num} scope{scope_num}, 
                # b: {num_round_2(b)}, R²: {num_round_2(R2)}")
                # Plot the fit line
                if y_axis != 'minor':
                    ax.plot(x, y_pred, color=colors[scope_num-1], linewidth=1)

                # Store fit results
                fit_results.append({
                    'mount': mount_num,
                    'scope': scope_num,
                    'slope': num_round_2(m),
                    'intercept': num_round_2(b),
                    'R2_score': num_round_2(R2),
                    'n_points': len(x)
                })
            # Scatter plot
            else:
                m = 0
                # For plotting all the points in the same symbol
                # ax.scatter(x, y, s=3, alpha=0.5, label=f'{scope_num} 
                # {num_round_2(m)} {num_round_2(R2,2)}')

            # for separate symbols for different days:
            for i in range(len(indexes) - 1):
                start = indexes[i]
                end = indexes[i+1]
                if 'R2' not in locals():
                    R2 = -1  # Make sure there is a value,to avoid errors 
                    # if missing
                ax.scatter(x[start:end], y[start:end],
                           marker=markers[i % len(markers)], s=15,
                           color=colors[scope_num-1],
                           # label only for first slice
                           label=
                           f'{scope_num} {num_round_2(m)} {num_round_2(R2, 2)}'
                           if i == 0 else None)

            lower_limit_mount = min(lower_limit_mount, lower_limit)
            upper_limit_mount = max(upper_limit_mount, upper_limit)
            if y_axis == 'minor':
                threshold = 3
                if len(x) > 0:
                    ax.plot([x.min(), x.max()], [threshold, threshold],
                            color='black', linewidth=2)
                lower_limit_mount = 1
                upper_limit_mount = 4
                temp = (mount_num, scope_num, len(x[y > threshold].tolist()))
                high_values.append(temp)

        # ax.set_ylim([lower_limit_mount-extra_y, upper_limit_mount+extra_y]) 
        #this is for absolute value of focus
        ax.set_ylim([34250, 35000])
        ax.set_title(f'Mount {mount_num}')
        ax.set_ylabel(y_axis)
        ax.grid(True, linestyle='--', alpha=0.3)
        if mount_num >= 8:
            ax.set_xlabel(x_axis)
        ax.legend(loc='upper left', fontsize=12)
    if not regular_fit and not constraints:
        title = title + ' slope=' + str(slope_fixed)

    fig.suptitle(title, fontsize=18, y=1.02)
    plt.tight_layout()
    filename = f'BestPos_vs_Temp_fits {earliest_time}-{latest_time}'
    label = str('focus')
    plot_saving(output_directory, filename, label)
    plt.show()

    # Now we export the Fit_results to an excel file and color the 
    #"bad fit columns
    #TODO findout why this code
    # write_append_cols_excel(fit_results,output_directory)

    df_high = pd.DataFrame(high_values, columns=['mount', 'scope', 'times'])
    # out = os.path.join(output_directory,(file_short + '_high.xlsx'))
    out = os.path.join(output_directory, (f'{date_range}' + '_high.xlsx'))
    df_high.to_excel(out, index=False)
    df_medians = pd.DataFrame(medians_list)
    file_name = os.path.join(
        output_directory, f'medians_BestPos_{date_range}.csv')
    df_medians.to_csv(file_name, index=False)
    print(f'saved medians to file:  medians_BestPos_{date_range}.csv')
    if plots.Focus_distribution_and_median:
        for row in data_for_distribution_plot:
            plot_distribution_if_outliers(row)

    return df_medians

def plot_distribution_if_outliers(input_tuple):
    """ This is used for plotting 40 (4 scopes x10 mounts) distributions
    of the BestPos from all the focuses carried out during the time of interest
    In general:
    Plots the distribution of a 1D data vector.

    Parameters:
    - data_vector (array-like): 1D data (e.g., df_scope[y_axis])
    - median (float): Precomputed median
    - std (float): Precomputed standard deviation
    - mount_num (int or str): Mount identifier for title
    - scope_num (int or str): Scope identifier for title
    """
    data_vector = input_tuple[0]
    median = input_tuple[1]
    std = input_tuple[2]
    mount_num = input_tuple[3]
    scope_num = input_tuple[4]
    date_range = input_tuple[5]
    # Drop NaNs if any
    data_vector = pd.Series(data_vector).dropna()
    n = len(data_vector)
    bins = max(int(np.ceil(np.log2(n)*3 + 1)), 10)
    plt.figure(figsize=(10, 6))

    # Plot histogram with frequency (default)
    counts, bins_edges, _ = plt.hist(data_vector, bins=bins, color='skyblue', edgecolor='black', alpha=0.6, label='Histogram')

    # Estimate KDE and scale it to match histogram frequencies
    kde = gaussian_kde(data_vector)
    x_vals = np.linspace(min(data_vector), max(data_vector), 1000)
    kde_vals = kde(x_vals) * n * (bins_edges[1] - bins_edges[0])  # Scale to counts

    # Plot KDE line
    plt.plot(x_vals, kde_vals, color='blue', linewidth=3, label='KDE')

    # Median line
    plt.axvline(median, color='green', linestyle='-',linewidth=3, label=f'Median: {median:.0f}')
    # Horizontal line showing 1 std extent (on X-axis)
    std_start = median - std
    std_end = median + std
    y_level = max(kde_vals) * 0.05  # Place line at 5% of peak KDE height
    plt.hlines(y=y_level, xmin=std_start, xmax=std_end, color='green', linewidth=3, label='±1 Std Dev')

    plt.title(f'BestPos Distribution - Mount {mount_num}, Scope {scope_num}, Std Dev = {std:.0f}')
    plt.xlabel('BestPos')
    plt.ylabel('Frequency')
    plt.legend()
    plt.tight_layout()
    plt.show()


def plot_tracking_results_1(tracking_results_groups,output_directory):
    '''This plots RA- and Dec-std [arcsec] vs. hour
       Probably shows windy conditions or other extremes
       and this is very useful'''
    
    for i, df in enumerate(tracking_results_groups):
        if df.empty:
            continue  # Skip empty groups    
        
        #extract the earliest and latest dates
        df["day_month"] = df["HH-MM-DD-mm"].str.split("_").str[-2:].str.join("_")
        df["date"] = pd.to_datetime(df["day_month"], format="%d_%m", errors="coerce").map(lambda x: x.replace(year=2000))
        earliest = df["date"].min().strftime("%d_%m")
        latest = df["date"].max().strftime("%d_%m")
            
        title = f'mount {i+1}: RA_std vs start_hour' + ' for: ' + earliest + ' --- ' + latest     
        
        plt.figure(figsize=(6, 4))
        plt.scatter(df['adjusted_hour'], df['RA_std'], marker='.', s=8, color='blue', alpha=0.7)
        plt.title(title)
        plt.xlabel('start_hour')
        plt.ylabel('RA_std [arcsec]')
        plt.ylim(0, 3)
        plt.grid(True)
        # Get the current Axes object
        ax = plt.gca()
        # Set major and minor ticks
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.16667))  # 6 minor per major
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))   # Optional for y-axis
        # Enable gridlines
        plt.grid(True, which='major', linewidth=0.8)
        plt.grid(True, which='minor', linestyle='--', linewidth=0.4, alpha=0.5)
        plt.tight_layout()
        filename = f'mount_{i+1}_RA_std'
        label = str('RA_std')
        plot_saving(output_directory, filename, label)
        plt.show()
    
    for i, df in enumerate(tracking_results_groups):
        if df.empty:
            continue  # Skip empty groups    
    
        plt.figure(figsize=(6, 4))
        plt.scatter(df['adjusted_hour'], df['Dec_std'], marker='.', s=8, color='red', alpha=0.7)
        plt.title(f'mount {i+1}: Dec_std vs start_hour')
        plt.xlabel('start_hour')
        plt.ylabel('Dec_std [arcsec]')
        plt.ylim(0, 0.3)
        plt.grid(True)
        # Get the current Axes object
        ax = plt.gca()
        # Set major and minor ticks
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.16667))  # 6 minor per major
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))   # Optional for y-axis
        # Enable gridlines
        plt.grid(True, which='major', linewidth=0.8)
        plt.grid(True, which='minor', linestyle='--', linewidth=0.4, alpha=0.5)
        plt.tight_layout()
        filename = f'mount_{i+1}_Dec_std'
        label = str('Dec_std')
        plot_saving(output_directory, filename, label)
        plt.show()
    
    #let's show how many rows for each mount
    #for i, df in enumerate(tracking_results_groups):
        #print(f"\nmount {i+1}: {df.shape[0]} rows")

def plot_FWHM(FWHM_groups,output_directory,showplots=False):
    '''Plots minor width [pix] without filtering on 10 separate pages, one for each mount.
    Each plot has 4 scatters for the 4 telescopes.
    The x-axis is the adjusted_hour plotted for each night from 17:30-02:30 UTC'''
    
    for i, df in enumerate(FWHM_groups):
        #let's show how many rows for each mount
        #print(f"\nmount {i+1}: {df.shape[0]} rows")

        if df.empty:
            continue  # Skip empty groups            
        # Get the unique scope values (expecting 4)
        scopes = sorted(df['scope'].unique())
        plt.ioff()
        # Start plotting
        plt.figure(figsize=(10, 6))
        
        for j, scope in enumerate(scopes):
            sub_df = df[df['scope'] == scope]
            plt.scatter(
                sub_df['adjusted_hour'],
                sub_df['minor'],
                color=colors[j],
                label=f'Scope: {scope}',
                marker='.',
                s=6,
                alpha=0.7
            )
        
        # Add labels, legend, etc.
        plt.xlabel('Adjusted Hour')
        plt.ylabel('FWHM [pix]')
        plt.title(f"mount {i+1}: Minor vs Adjusted Hour")
        plt.legend()
        # Get the current Axes object
        ax = plt.gca()
        # Set major and minor ticks
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.16667))  # 6 minor per major
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))   # Optional for y-axis
        # Enable gridlines
        plt.grid(True, which='major', linewidth=0.8)
        plt.grid(True, which='minor', linestyle='--', linewidth=0.4, alpha=0.5)
        plt.ylim(1, 7)
        plt.xlim(17.5, 26.5)
        plt.tight_layout()
        filename = f'mount_{i+1}_FWHM'
        label = str('FWHM_no_averaging')
        plot_saving(output_directory, filename, label)
        if showplots:
            plt.show()


def plot_tracking_results_2(tracking_results_groups,output_directory):
    '''10 scatter plots for the different mounts showing both RA and Dec
    as a function of adjusted_hour (17:00-02:30)'''
    
    for i, df in enumerate(tracking_results_groups):
        if df.empty:
            continue  # Skip empty groups    
        
        plt.figure(figsize=(6, 4))
        plt.scatter(df['adjusted_hour'], df['RA_median'], marker='.', s=8, label = 'RA', color='blue', alpha=0.7)
        plt.scatter(df['adjusted_hour'], df['Dec_median'],marker='+', s=6, label = 'Dec', color='red', alpha=0.7)
        
        #extract the earliest and latest dates
        df["day_month"] = df["HH-MM-DD-mm"].str.split("_").str[-2:].str.join("_")
        df["date"] = pd.to_datetime(df["day_month"], format="%d_%m", errors="coerce").map(lambda x: x.replace(year=2000))
        earliest = df["date"].min().strftime("%d_%m")
        latest = df["date"].max().strftime("%d_%m")
            
        title = f'mount {i+1}: RA/Dec vs start_hour' + ' for: ' + earliest + ' --- ' + latest     
        plt.title(title)
        plt.xlabel('start_hour')
        plt.ylabel('RA , Dec [deg]')
        plt.ylim(-180, 180)
        plt.xlim(17., 26.5)
        plt.legend()
        # Get the current Axes object
        ax = plt.gca()
        # Set major and minor ticks
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.16667))  # 6 minor per major
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))   # Optional for y-axis
        # Enable gridlines
        plt.grid(True, which='major', linewidth=0.8)
        plt.grid(True, which='minor', linestyle='--', linewidth=0.4, alpha=0.5)
        plt.tight_layout()
        filename = f'mount_{i+1}_{earliest}-{latest}'
        label = str('RA_Dec_hourly')
        plot_saving(output_directory, filename, label)
        plt.show()
        
def filter_only_noisy_std(tracking_results, threshold_hour=19):    
    '''Check for Noisy RA and Dec by looking only at t > threshold_hour(default = 19:00)
    and RA.std > 1.5 arcsec and Dec.std>0.2 arcsec. The number of "Noisy" observations
    is printed for each mount and the Noisy df's are returned. This can  show if
    the mounts are unable to track accurately. It DOES NOT show the scope flapping in the wind'''    
    
    print('\nfilter_only_noisy_std function:')
    RA_std_threshold = 1.5  # Set your threshold here [arcsec]
    RA_filtered_results = []
    for df in tracking_results:
        if df.empty:
            RA_filtered_results.append(df.copy())
            continue
    
        # Apply filters: std > threshold + hour > threshold_hour
        mask = (df['RA_std'] > RA_std_threshold) & (df['adjusted_hour'] > threshold_hour)
    
        filtered_df = df[mask].copy()
        RA_filtered_results.append(filtered_df)
    print('\nthreshold RA_std=  ', RA_std_threshold, ' leaves ', len(filtered_df), 'events')
    for i, df in enumerate(RA_filtered_results):
        count = len(df)
        print(f"mount {i+1}: {count} high_RA_std after {threshold_hour}")
    
    
    Dec_std_threshold = 0.2  # Set your threshold here [arcsec]
    Dec_filtered_results = []
    for df in tracking_results:
        if df.empty:
            Dec_filtered_results.append(df)
            continue
        
        # Apply filters: std > threshold + hour > threshold_hour
        mask = (df['Dec_std'] > Dec_std_threshold) & (df['adjusted_hour'] > threshold_hour) 
    
        filtered_df = df[mask].copy()
        Dec_filtered_results.append(filtered_df)
    print('\nthreshold Dec_std=  ', Dec_std_threshold, ' leaves ', len(filtered_df), 'events')
    for i, df in enumerate(Dec_filtered_results):
        count = len(df)   
        print(f"mount {i+1}: {count} high_Dec_std after {threshold_hour}")
        
    return RA_filtered_results, Dec_filtered_results

def insert_FWHM_into_tracking_results_groups(tracking_results_groups, FWHM_groups):
    '''tracking_results_group was previously created with a df for each mount. 
    Here we append the columns by adding for each scope a column of FWHM during
    the relevant observation window. Each inserted FWHM is the median during the
    observation window. Both minor and major for each of the 4 scopes are added'''
    
    updated_tracking_groups = []

    for FWHM_df, target_df in zip(FWHM_groups, tracking_results_groups):
        target_df = target_df.copy()
    
        # Prepare empty columns for each scope
        for scope_id in range(1, 5):
            target_df[f'minor_{scope_id}'] = np.nan
    
        # Go through each row in the target_df
        for idx, row in target_df.iterrows():
            start = row['start']
            finish = row['finish']
    
            # Filter source_df rows within time window
            mask = (FWHM_df['TimeStarted'] >= start) & (FWHM_df['TimeStarted'] <= finish)
            matched_rows = FWHM_df[mask]
    
            # For each scope, average 'minor' values in matched_rows
            for scope_id in range(1, 5):
                scope_rows = matched_rows[matched_rows['scope'] == scope_id]
                if not scope_rows.empty:
                    avg_minor = scope_rows['minor'].median()
                    avg_major = scope_rows['major'].median()
                    target_df.at[idx, f'minor_{scope_id}'] = avg_minor
                    target_df.at[idx, f'major_{scope_id}'] = avg_major

        updated_tracking_groups.append(target_df)
    return updated_tracking_groups

def plot_tracking_results_3(tracking_results_groups, focus_groups,output_directory):
    '''10 scatter plots for the different mounts showing FWHM (left) and RA and Dec (right)
    as a function of adjusted_hour (17:30-02:30) and Vertical lines for Focuses'''
    for i, df in enumerate(tracking_results_groups):
        if df.empty:
            continue  # Skip empty groups    
        #Need the start time range for clipping the focus data
        start_time = df['start'].min()
        end_time = df['start'].max()
        end = df['HH-MM-DD-mm'].min()[-5:] #for the title we want only DD-mm
        start = df['HH-MM-DD-mm'].max()[-5:]
        
        '''This part plots the RA and Dec on the right axis as ax2
        and a scatter plot of minor1-4 on ax'''
        # Create figure and base axes
        fig, ax = plt.subplots(figsize=(6, 4))
        ax2 = ax.twinx()
        
        # === Plot 'RA_mean' and 'Dec_mean' on the right y-axis ===
        #ax2.scatter(df['adjusted_hour'], df['RA_median'], label='RA', color='orange', marker='x', s=5,  alpha=0.7)
        #ax2.scatter(df['adjusted_hour'], df['Dec_median'], label='Dec', color='black', marker='x', s=5,  alpha=0.7)
        
        # === Plot 'RA_mean' and 'Dec_mean' on the right y-axis ===
        ax2.scatter(df['adjusted_hour'], df['Az_median'], label='Az', color='orange', marker='x', s=5,  alpha=0.7)
        ax2.scatter(df['adjusted_hour'], df['Alt_median'], label='Alt', color='black', marker='x', s=5,  alpha=0.7)
        
        # === Plot 'minor_1' to 'minor_4' on the left y-axis ===
        
        for j in range(1, 5):
            col = f'minor_{j}'
            if col in df.columns:
                ax.scatter(df['adjusted_hour'], df[col], s=6, color=colors[j - 1], alpha=0.7, label=col)
        
        # Axis labels and title
        ax.set_xlabel('Start Hour')
        ax.set_ylabel('FWHM [pix]')
        ax2.set_ylabel('RA , Dec or Az, Alt [deg]')
        plt.title(f'Mount {i + 1}: FWHM + RA/Dec {start}-{end}')
        
        # Tick settings
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(1 / 6))
        ax2.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))
        
        # Grid
        ax.grid(True, which='major', linewidth=0.8)
        ax.grid(True, which='minor', linestyle='--', linewidth=0.4, alpha=0.5)
        ax2.grid(False)
        
        # Legends: combine from both axes
        handles1, labels1 = ax.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        ax2.legend(handles1 + handles2, labels1 + labels2, loc='upper left')
        ax2.set_ylim(-320, 100)
        ax.set_ylim(1, 6)
        # Layout
        plt.tight_layout()
        # Lock the current x-axis limits
        x_min, x_max = ax.get_xlim()
        # Assume `event_groups` is a list of DataFrames matching each tracking_results_group
        
        # add vertical black lines for the focusing sessions
        if i < len(focus_groups):
            focus_df = focus_groups[i]
            focus_df = focus_df[(focus_df['TimeStarted'] >= start_time) & (focus_df['TimeStarted'] <= end_time)]
            if not focus_df.empty and 'adjusted_hour' in focus_df.columns:
                last_plotted = -float('inf')  # Track last plotted time
                for event_time in sorted(focus_df['adjusted_hour']):
                    if event_time - last_plotted >= 0.1:
                        ax.axvline(x=event_time, color='black', linewidth=0.8, linestyle='-', alpha=0.8)
                        last_plotted = event_time

        #ax.set_xlim(x_min, x_max)
        ax.set_xlim(17.5, 26.5)
        filename = f'mount_{i+1}_{start}-{end}'
        label = str('FWHM_RA_Dec')
        plot_saving(output_directory, filename, label)
        plt.show()

def fast_line_count(filename):
    '''run a linux command to get the number of lines in the csv filename
    this is used for calculating the number of lines to read (that correspond to
    a fraction requested by user'''
    result = subprocess.run(['wc', '-l', filename], capture_output=True, text=True)
    parts = result.stdout.strip().split()
    return int(parts[0]) if parts else 0

def add_Alt_to_focus(df_focus, df_Alt):
    '''For each row of focusing, look into the Alt df and find the average Alt at which that focus was performed'''
    # Make a copy to avoid modifying the original
    df_focus = df_focus.copy()
    
    alts = []  # will store the computed alt values
    for idx, row in df_focus.iterrows():
        # idx is something like "04.3"
        mount_scope = str(idx)
        # Extract the mount (the part before the dot)
        if "." in mount_scope:
            mount_str = mount_scope.split(".")[0]
        else:
            # fallback if the index isn't in mount.scope format
            mount_str = mount_scope
        # Time window
        t_start = row["TimeStarted"]
        t_end   = row["TimeEnded"]
        # Filter df_Alt by mount and time range
        # 1) match mount
        alt_subset = df_Alt[df_Alt.index.str.startswith(mount_str + ".")]
        # 2) within time range
        alt_subset = alt_subset[ (alt_subset["TimeStarted"] >= t_start) &
                                 (alt_subset["TimeStarted"] <= t_end)    ]
        if len(alt_subset) == 0:
            alts.append(np.nan) # No match
        else:
            # Average of 'value' column, rounded to 1 decimal
            alt_subset['value'] = pd.to_numeric(alt_subset['value'], errors='coerce')
            mean_val = alt_subset["value"].mean()
            alts.append(round(mean_val, 1))
    df_focus["Alt"] = alts
    return df_focus



def get_focus_windows(df, t_min=None, t_max=None):
    """ Produce a list of focus windows with start, finish, obs_duration.
    The focus windows are actually observation windows between adjacent focus sessions
    of the same scope. A sub-list is returned whose duration is >t_min and <t_max.
    Also, in order to remove windows to the next day, we add an artifical focus at 03:00
    for all the scopes
    
    Adds Temp_list : list of Temperature values from original df that fall
    within the start_obs - finish_obs range, per mount and scope. """
    
    out = df.copy()
    
    int_part, frac_part = np.divmod(out.index.to_numpy(dtype=float), 1)
    # Assign mount (0–9) and scope (0–3)
    out['mount'] = (int_part).astype(int)
    out['scope'] = np.round(frac_part * 10).astype(int)
    #out['int_index'] =  int_part.astype(int) 
    # sort by mount, scope, and TimeStarted
    out = out.sort_values(["mount", "scope", "TimeStarted"])

    # start = current TimeEnded
    out["start_obs"] = out["TimeEnded"]

    # finish = next TimeStarted within (mount, scope)
    out["finish_obs"] = out.groupby(["mount", "scope"])["TimeStarted"].shift(-1)
    # in order to also take the last observations after the last focusing session, we add the following
    out["finish_obs"] = out["finish_obs"].fillna(out["TimeStarted"] + 0.5)  # if nan, add 0.5 days

    # duration in seconds
    out["obs_duration"] = round((out["finish_obs"] - out["start_obs"]) * 86400)
    #remove empty obs duration if any duration = nan
    out = out[~out["obs_duration"].isna()].copy()    
    # Add Temp_list: all Temperature values from original df within the window
    temp_lists = []
    for idx, row in out.iterrows():
        mask = (
            (out["mount"] == row["mount"]) &
            (out["scope"] == row["scope"]) &
            (out["TimeStarted"] >= row["start_obs"]) &
            (out["TimeStarted"] <= row["finish_obs"])
        )
        temp_lists.append(out.loc[mask, "Temperature"].tolist())

    out["Temp_list"] = temp_lists
    out["BestPos"] = pd.to_numeric(out["BestPos"], errors="coerce").round(0).astype("Int64").astype(str)

    # put new columns first
    first = ["start_obs", "finish_obs", "obs_duration", "Temp_list"]
    out = out[first + [c for c in out.columns if c not in first]]

    # filter if requested
    if t_min is not None or t_max is not None:
        mask = pd.Series(True, index=out.index)
        if t_min is not None:
            mask &= out["obs_duration"] > t_min
        if t_max is not None:
            mask &= out["obs_duration"] < t_max 
        out = out[mask]

    return out


def append_temperature_deltas(focus_times, max_dT=0.5):
    '''Adds a column called delta_Temperature that has the delta_T between adjacent rows '''
    # Make a copy to avoid modifying original inplace
    df = focus_times.copy()
    # sort to ensure proper order
    df = df.sort_values(["mount", "scope", "start_obs"])
    # compute delta temperature directly (shift -1 within groups)
    delta = (df.groupby(["mount", "scope"])["Temperature"].shift(-1) - df["Temperature"]) 
    # apply threshold filter
    df["delta_Temperature"] = delta.where(delta < max_dT, np.nan)
    return df    


def get_minor_series(focus_times, tracking_results_groups):
    '''For each of the focus_times, extracts the minor, major and Alt and Az from
    tracking_results_groups. 2 new df's are created: combined and jumps
    combined is then plotted using plot_KDE
    delta_Minor and delta_Major are the max(FWHM)-min(FWHM) of the observations between adjacent focusing
    sessions. Typically, if these are 1 hr apart they include ~9 observations'''
    
    results = []
    jumps   = []
    print(f'Found {len(focus_times)} focus times')
    for _, row in focus_times.iterrows():
        mount = int(row["mount"])
        scope = int(row["scope"])
        start_obs = row["start_obs"]
        finish_obs = row["finish_obs"]

        # pick the correct tracking_results df
        tr_df = tracking_results_groups[mount - 1]

        # filter condition
        mask = (tr_df["start"] > start_obs) & (tr_df["start"] < finish_obs)
        n_rows = mask.sum()
        
        # select the scope-specific column
        col_minor = f"minor_{scope}"
        col_major = f"major_{scope}"
        if col_minor not in tr_df.columns:
            continue  # skip if missing
        # extract lists
        #minor_list = tr_df.loc[mask, col_minor].tolist()
        minor_list = [round(x, 3) if pd.notna(x) else x for x in tr_df.loc[mask, col_minor].tolist()]
        if col_major not in tr_df.columns:
            continue  # skip if missing
        # extract lists
        #major_list = tr_df.loc[mask, col_minor].tolist()
        major_list = [round(x, 3) if pd.notna(x) else x for x in tr_df.loc[mask, col_major].tolist()]
        
        az_list    = tr_df.loc[mask, "Az_median"].tolist()
        alt_list   = tr_df.loc[mask, "Alt_median"].tolist()
        #Temp_list  = tr_df.loc[mask, "Temperature"].tolist()
        Temp_list = [round(x, 2) if pd.notna(x) else x for x in tr_df.loc[mask, "Temperature"].tolist()]
        # compute delta_minor if possible
        delta_minor = max(minor_list) - min(minor_list) if minor_list else None
        delta_major = max(major_list) - min(major_list) if major_list else None
        # df_dT
        delta_Temp = row["delta_Temperature"]
        if delta_Temp not in (None, 0) and delta_minor not in (None, 0):
            dMinor_dT = (delta_minor / delta_Temp) * (-1)
        else:
            dMinor_dT = None
        
        #Generate a Dictionary of jumps in minor where the value is a 4 element tuple
        # 1. change in Az 2. change in Az 3. mount.scope  4. HH-MM-DD-mm (time)
        vals_minor = tr_df.loc[mask, col_minor].dropna().tolist()
        vals_az = tr_df.loc[mask, "Az_median"].dropna().tolist()
        vals_alt = tr_df.loc[mask, "Alt_median"].dropna().tolist()
        times = tr_df.loc[mask, "HH-MM-DD-mm"].tolist()
        
        jump_dic = {}
        
        for i in range(len(vals_minor) - 1):
            jump_minor = round(vals_minor[i+1] - vals_minor[i], 3)
            if abs(jump_minor) > 0.15:
                jump_az = round( (vals_az[i+1] - vals_az[i] + 180) % 360 -180, 3) #make sure the diff remains in the -180 - +180 range
                jump_alt = round(vals_alt[i+1] - vals_alt[i], 3)
                jump_sin_alt = float(np.sin(np.deg2rad(vals_alt[i+1])) - np.sin(np.deg2rad(vals_alt[i])))
                jump_distance = round( angular_distance(vals_alt[i+1], vals_az[i+1],
                                vals_alt[i], vals_alt[i]) , 3)
                jump_ini_alt = round(vals_alt[i], 3)
                timestamp = times[i+1]  # use the later time for this step
                mount_scope = f"{mount}.{scope}"  # assumes variables mount and scope exist
                jump_dic[jump_minor] = (jump_az, jump_alt, jump_sin_alt, jump_distance, jump_ini_alt, timestamp, mount_scope)    
                
        
        results.append({
            "mount": mount,
            "scope": scope,
            "start_obs": start_obs,
            "finish_obs": finish_obs,
            "HH-MM-DD-mm": row['HH-MM-DD-mm'],
            "FocusPosition": row['BestPos'],
            "Temperature": row["Temperature"],
            "delta_Temperature": row["delta_Temperature"],
            "delta_Minor": delta_minor,
            "delta_Major": delta_major,
            "dMinor_dT": dMinor_dT,
            "Temp_list": Temp_list,
            "minor_list": minor_list,
            "major_list": major_list,
            "az_list": az_list,
            "alt_list": alt_list,
            "n_rows": n_rows  # keep row count for inspection
        })
        if "jump_dic" in locals() and jump_dic:   # check that jump_dic exists and is not empty
            for key, value in jump_dic.items():
                jumps.append({
                "jump_minor": key,          # the dict key
                "jump_az": value[0],        # first element of tuple
                "jump_alt": value[1],       # second element of tuple
                "jump_sin_alt": value[2],    # 3rd element of tuple
                "jump_distance": value[3],  # 4rd element of tuple
                "jump_ini_alt": value[4],      # 5th element of tuple
                "jump_time": value[5],      # 6th element of tuple
                "jump_mountscope": value[6] # 7th element of tuple
            })

    return pd.DataFrame(results), pd.DataFrame(jumps)


def plot_KDE(df, mount, column, output_directory, scopes=[1,2,3,4], bw_adjust=1):
    """ Plot KDE curves for a given mount and column, for the different scopes.
    Parameters:
    - df : DataFrame
    - mount : int, mount number to filter
    - column : str, column name to plot
    - scopes : list of int, scopes to include
    - bw_adjust : float, smoothing parameter for KDE """
    
    plt.figure(figsize=(10,6))

    for scope, color in zip(scopes, colors):
        subset = df[(df["mount"] == mount) & (df["scope"] == scope)][column].dropna()
        if len(subset) == 0:
            continue
        sns.kdeplot(subset, color=color, lw=2, label=f"scope {scope}", bw_adjust=bw_adjust, warn_singular=False)

    plt.title(f"KDE of {column} for mount={mount}, [max(FWHM)-min(FWHM)] between adjacent focuses")
    plt.xlabel(column)
    plt.xlim(-2,7)
    plt.ylabel("Density")
    plt.grid(True, linestyle="--", alpha=0.7)
    handles, labels = plt.gca().get_legend_handles_labels()
    if labels:
        plt.legend() # THE IF helps avoid warning when no data available
    plt.tight_layout()
    filename = f'mount_{i+1}_{col_name}'
    label = f'histograms {col_name}'
    plot_saving(output_directory, filename, label)
    plt.show()
  

def smoothness_score(trace):
    '''A helper function to return the std of the fluctuations of a trace'''
    diffs = np.diff(trace)
    return np.std(diffs)

def smoother_df(df1, df2, col="value"): # meant for Temp traces, helps to choose the better one
    """Return the DataFrame with the smoother trace in given column. 
    This is used for selecting Temp_1 or Temp_2 as the more reliable
    temperature for each mount"""
    df1['value'] = pd.to_numeric(df1['value'], errors='coerce')
    df2['value'] = pd.to_numeric(df2['value'], errors='coerce')
    s1 = np.std(np.diff(df1[col].dropna()))
    s2 = np.std(np.diff(df2[col].dropna()))
    return df1 if s1 < s2 else df2
    

def add_temp_medians(tracking_results_groups, df_Temp_groups):
    """ For each df in tracking_results_groups:
      - get start, finish from each row
      - lookup rows in df_Temp_groups[i] whose TimeStarted falls in that window
      - compute median Temperature for each scope
      - add columns Temp_1...Temp_4
    Returns a new list of updated tracking_results_groups. """
    
    updated_groups = []

    for i, track_df in enumerate(tracking_results_groups):
        Temp_df = df_Temp_groups[i].copy()
        track_copy = track_df.copy()

        medians = []
        for _, row in track_copy.iterrows():
            start, finish = row["start"], row["finish"]
            # mask rows within the time window AND for this scope           
            mask = ((Temp_df["TimeStarted"] >= start) &
                    (Temp_df["TimeStarted"] <= finish) )                     
            values = Temp_df.loc[mask, "value"]
            # compute median or NaN
            if not values.empty:
                median_val = np.round(values.median(),4)
            else:
                median_val = np.nan
            medians.append(median_val)
        # add the new column
        track_copy["Temperature"] = medians
        updated_groups.append(track_copy)

    return updated_groups


def plot_jumps(jumps,output_directory):
    '''Plots the jumps [pix] in 4 bins: 0.3-0.7, 0.7-1.4, 1.4-2, >2  for each scope'''
    
    freq_jumps = (jumps["jump_mountscope"].value_counts().rename("freq").rename_axis("jump_mountscope").reset_index())
    freq_jumps = freq_jumps.sort_values(by="jump_mountscope", key=lambda col: col.astype(str).map(lambda x: tuple(map(int, x.split("."))))).set_index("jump_mountscope")
    bins = [0.3, 0.7, 1.4, 2, float("inf")]
    labels = ["0.3–0.7", "0.7–1.4", "1.4–2", ">2"]
    jumps["bin"] = pd.cut(jumps["jump_minor"].abs(), bins=bins, labels=labels, right=True)
    counts = (jumps.groupby(["jump_mountscope", "bin"], observed=True).size().unstack(fill_value=0))
    # Ensure index is sorted numerically by mount then scope
    counts_sorted = counts.copy()
    counts_sorted["mount"] = counts_sorted.index.map(lambda x: int(x.split(".")[0]))
    counts_sorted["scope"] = counts_sorted.index.map(lambda x: int(x.split(".")[1]))
    counts_sorted = counts_sorted.sort_values(["mount", "scope"])
    
    markers = {"0.3–0.7": ".", "0.7–1.4": "o", "1.4–2": "x", ">2": "D"}  # circle, square, diamond
    plt.figure(figsize=(10, 6))
    
    for scope, row in counts_sorted.iterrows():
        mount, sc = scope.split(".")
        color = colors_mounts.get(mount, "black")
        for b in labels:
            if b in row and not pd.isna(row[b]):
                plt.scatter(scope, row[b], marker=markers[b], color=color, s=60)
    
    legend_handles = [
        mlines.Line2D([], [], color="black", marker=".", linestyle="None", label="0.3–0.7"),
        mlines.Line2D([], [], color="black", marker="o", linestyle="None", label="0.7–1.4"),
        mlines.Line2D([], [], color="black", marker="x", linestyle="None", label="1.4–2"),
        mlines.Line2D([], [], color="black", marker="D", linestyle="None", label=">2"),    ]   
    
    #extract the earliest and latest dates
    jumps["day_month"] = jumps["jump_time"].str.split("_").str[-2:].str.join("_")
    jumps["date"] = pd.to_datetime(jumps["day_month"], format="%d_%m", errors="coerce").map(lambda x: x.replace(year=2000))
    earliest = jumps["date"].min().strftime("%d_%m")
    latest = jumps["date"].max().strftime("%d_%m")        
    title = 'Jump counts per telescope  for: ' + earliest + ' --- ' + latest     
    
    plt.legend(handles=legend_handles, loc="best")
    plt.ylabel("Jump counts")
    plt.xlabel("Telescope (mount.scope)")
    plt.title(title)
    plt.xticks(rotation=90)
    plt.grid(True)
    plt.tight_layout()
    filename = 'jumps_per_scope'
    label = 'jumps'
    plot_saving(output_directory, filename, label)
    plt.show()


def plot_jumps_2(jumps,output_directory):
    '''Plots 3-sub plots for demonstrating the correlation of the jumps (which
    are on the x-axis) with the az_jump, alt_jump and sin(alt_jump).'''
    fig, axes = plt.subplots(nrows=3, ncols=1, figsize=(8, 10), sharex=True)
    earliest = jumps["date"].min().strftime("%d_%m")
    latest = jumps["date"].max().strftime("%d_%m")        
    title = 'Jump relationships for: ' + earliest + ' --- ' + latest 
    # First subplot: jump_minor vs jump_az
    axes[0].scatter(jumps["jump_minor"], jumps["jump_az"], s=5, c="tab:blue")
    axes[0].set_ylabel("jump_az")
    axes[0].set_title(title)

    # Second subplot: jump_minor vs jump_alt
    axes[1].scatter(jumps["jump_minor"], jumps["jump_alt"], s=5, c="tab:orange")
    axes[1].set_ylabel("jump_alt")

    # Third subplot: jump_minor vs jump_distance
    axes[2].scatter(jumps["jump_minor"], jumps["jump_sin_alt"], s=5,c="tab:green")
    axes[2].set_ylabel("jump_sin_alt")
    axes[2].set_xlabel("jump_minor")

    plt.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    filename = f'3-sub_plots_jumps {earliest}-{latest}'
    label = str('jumps')
    plot_saving(output_directory, filename, label)
    plt.show()


def angular_distance(alt1_deg, az1_deg, alt2_deg, az2_deg):
    '''Calculates the angular distance between two (Alt,Az) coordinates and returns it in degrees'''
    
    # convert to radians
    alt1 = math.radians(alt1_deg)
    alt2 = math.radians(alt2_deg)
    az1 = math.radians(az1_deg)
    az2 = math.radians(az2_deg)

    cos_d = math.sin(alt1)*math.sin(alt2) + math.cos(alt1)*math.cos(alt2)*math.cos(az1 - az2)
    cos_d = max(-1, min(1, cos_d))  # clamp for safety
    d = math.acos(cos_d)
    return math.degrees(d)


def mount_scope_medians(FWHM_groups,output_directory):
    """  Take a list of DataFrames and calculate median 'minor' for each scope in each df.
    Returns a single DataFrame with columns:
      index (1..N), mount, scope, avg_minor    
   Optionally merges saved data with the current data   
      """
      
    rows = []   
    for df in FWHM_groups:
        if "mount" not in df.columns or "scope" not in df.columns or "minor" not in df.columns:
            continue  # skip if required columns missing
        
        temp = df['HH-MM-DD-mm'].dropna().astype(str).max()
        parts = str(temp).split('_')
        earliest = ''.join(parts[-2:]) 
        temp   = df['HH-MM-DD-mm'].dropna().astype(str).min()
        parts = str(temp).split('_')
        latest = ''.join(parts[-2:]) 
        #I don't know why, but this kept on giving an error
        #earliest = df["time"].min().strftime("%d_%m")
        #latest = df["time"].max().strftime("%d_%m") 
        grouped = df.groupby(["mount", "scope"])["minor"].median().round(3).reset_index(name="avg_minor")
        rows.append(grouped)
    
    # Combine all
    all_df = pd.concat(rows, ignore_index=True)
    # Ensure mount and scope are strings
    all_df["mount_str"] = all_df["mount"].astype(int).astype(str)
    all_df["scope_str"] = all_df["scope"].astype(int).astype(str)
    # Format mount with leading zero if needed
    all_df["mount_str"] = all_df["mount_str"].str.zfill(2)   
    all_df["mount_scope"] = all_df["mount_str"] + "." + all_df["scope_str"]
    #TODO take out the plot to a function
    # def plotXXXXX(all_df:pd.DataFrame)
    # plt.figure(figsize=(10, 5))
    # #TODO  generalize to num of mounts
    # for _, row in all_df.iterrows():
    #     mount = str(int(row["mount"]))  # e.g., "1","2",..."10"
    #     label = row["mount_scope"]
    #     value = row["avg_minor"]
    #     plt.scatter(label, value, color=colors_mounts.get(mount, "black"), s=80 )       
    # title = 'Median FWHM vs Mount.Scope for ' + earliest + ' --- ' + latest 
    # plt.xlabel("Mount.Scope")
    # plt.ylabel("Median Minor (avg_minor)")
    # plt.title(title)
    # plt.grid(True, linestyle="--", alpha=0.4)
    # plt.xticks(rotation=90)   
    # plt.tight_layout()
    # filename = f'median_FWHM_all_scopes {earliest}-{latest}'
    # label = str('median_FWHM')
    # plot_saving(output_directory, filename, label)
    # plt.show()

    # Sort by mount, then scope
    all_df = all_df.sort_values(["mount", "scope"]).reset_index(drop=True)
    
    # Add running index starting from 1
    all_df.index = all_df.index + 1
    all_df.index.name = "index"
    if plot_Focus_temperature_slope: # can only merge if indeed ran focus_temperature fit
        file = os.path.join(output_directory,(file_short + '_Fit_results.xlsx'))
        df_R2 = pd.read_excel(file, usecols=["mount", "scope","slope", "R2_score"])
        # Merge on (mount, scope)
        all_df = pd.merge(
            all_df,
            df_R2,
            on=["mount", "scope"],
            how="left"   # keep all rows from all_df
        )
    return all_df


def plot_alt_vs_hour(focus_groups:list,output_directory:str):
    '''plots the Alt at which each mount performed focus at the different hours'''
    
    fig, ax = plt.subplots(figsize=(10, 6))  
    for mount_idx, df_focus in enumerate(focus_groups, start=1):
        # Skip empty or missing required columns
        if df_focus.empty or "Alt" not in df_focus.columns or "scope" not in df_focus.columns:
            continue
        earliest = df_focus["time"].min().strftime("%d_%m")
        latest = df_focus["time"].max().strftime("%d_%m") 
        # Get the lowest available scope numerically
        scopes = df_focus["scope"].dropna().unique()
        if len(scopes) == 0:
            continue

        selected_scope = np.sort(scopes.astype(int))[0]

        # Filter by that scope
        df_scope = df_focus[df_focus["scope"] == selected_scope]

        if df_scope.empty:
            continue
        
        # Drop NaN values in Alt or adjusted_hour
        df_scope = df_scope.dropna(subset=["Alt", "adjusted_hour"])
        if df_scope.empty:
            continue

        # Sort by x so the line connects in time order
        df_scope = df_scope.sort_values("adjusted_hour")
        
        ax.plot(
            df_scope["adjusted_hour"],
            df_scope["Alt"],
            marker="o",
            markersize=10,
            linewidth=0.5,
            label=f"Mount {mount_idx}",
            color=colors_mounts.get(str(mount_idx), "black")
        )

    ax.set_xlabel("Adjusted Hour [H]")
    ax.set_ylabel("Alt [deg]")
    title = 'Alt during Focusing at different Hours ' + earliest + ' --- ' + latest 
    ax.set_title(title)
    ax.grid(True)
    ax.legend(loc="best")
    plt.tight_layout()
    filename = f'Alt_at_which_focused {earliest}-{latest}'
    label = str('focus')
    plot_saving(output_directory, filename, label)
    plt.show()
